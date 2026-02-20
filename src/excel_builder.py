import io
import os

import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PilImage

from src.config import TEMP_DIR, THUMB_HEIGHT, THUMB_WIDTH

COLUMNS = [
    ("Thumbnail", 14),
    ("Name", 28),
    ("Set", 24),
    ("Faction", 14),
    ("Rarity", 12),
    ("Type", 14),
    ("Cost", 8),
    ("Recall Cost", 12),
    ("Mountain", 10),
    ("Ocean", 10),
    ("Forest", 10),
    ("Quantity", 10),
    ("Full Image", 14),
]

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical="center", wrap_text=False)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
ALT_ROW_FILL = PatternFill(start_color="F2F6FC", end_color="F2F6FC", fill_type="solid")

ROW_HEIGHT_PX = THUMB_HEIGHT + 4
ROW_HEIGHT_PT = ROW_HEIGHT_PX * 0.75


def _download_thumbnail(image_url: str, reference: str) -> str | None:
    """Download card image and save a resized thumbnail. Returns path or None."""
    if not image_url:
        return None

    os.makedirs(TEMP_DIR, exist_ok=True)
    thumb_path = os.path.join(TEMP_DIR, f"{reference}.png")

    if os.path.exists(thumb_path):
        return thumb_path

    try:
        resp = requests.get(image_url, timeout=30)
        resp.raise_for_status()
    except requests.RequestException:
        return None

    try:
        img = PilImage.open(io.BytesIO(resp.content))
        img.thumbnail((THUMB_WIDTH, THUMB_HEIGHT), PilImage.LANCZOS)
        img.save(thumb_path, "PNG")
        return thumb_path
    except Exception:
        return None


def build_catalogue(cards: list[dict], output_path: str) -> None:
    """Build an Excel catalogue from the flat card list."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Altered Cards"

    # --- header row ---
    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    # --- data rows ---
    total = len(cards)
    for idx, card in enumerate(cards):
        row = idx + 2
        if idx % 50 == 0:
            print(f"  Building row {row - 1}/{total} ...")

        # thumbnail (column A)
        thumb_path = _download_thumbnail(card["image_url"], card["reference"])
        if thumb_path:
            try:
                img = XlImage(thumb_path)
                img.width = THUMB_WIDTH
                img.height = THUMB_HEIGHT
                ws.add_image(img, f"A{row}")
            except Exception:
                pass

        values = [
            "",  # A: thumbnail placeholder
            card["name"],
            card["card_set"],
            card["faction"],
            card["rarity"],
            card["card_type"],
            _to_int(card["main_cost"]),
            _to_int(card["recall_cost"]),
            _to_int(card["mountain_power"]),
            _to_int(card["ocean_power"]),
            _to_int(card["forest_power"]),
            0,  # Quantity
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER
            if idx % 2 == 1:
                cell.fill = ALT_ROW_FILL

        # Full Image hyperlink (column M)
        link_cell = ws.cell(row=row, column=13)
        if card["image_url"]:
            link_cell.hyperlink = card["image_url"]
            link_cell.value = "View Full"
            link_cell.font = Font(color="0563C1", underline="single")
        link_cell.alignment = CELL_ALIGNMENT
        link_cell.border = THIN_BORDER
        if idx % 2 == 1:
            link_cell.fill = ALT_ROW_FILL

        ws.row_dimensions[row].height = ROW_HEIGHT_PT

    # --- autofilter (columns B through K: Name..Forest) ---
    last_row = len(cards) + 1
    last_col_letter = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

    # --- freeze header row ---
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"Saved catalogue to {output_path}")


def _to_int(val: str):
    """Convert numeric string to int, returning empty string for blanks."""
    if val is None or val == "":
        return ""
    try:
        return int(val)
    except (ValueError, TypeError):
        return val
